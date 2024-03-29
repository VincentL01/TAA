import tkinter
import tkinter.messagebox
import tkinter.ttk as ttk
import customtkinter

import json
from pathlib import Path
import shutil
import os
import pandas as pd
import time
import threading

import openpyxl

import logging
from colorlog import ColoredFormatter

from Libs.autoanalyzer import autoanalyzer
from Libs.importvideos import VideoAdd
from Libs.misc import get_static_dir, check_trajectories_dir, load_raw_df, get_treatment_name_from_char, get_treatment_char_from_name, init_history
from Libs.constants import CONSTANTS
from Libs import TESTS_LIST, ORDINALS, CHARS
from Libs.customwidgets import Selector


customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"


#[TODO] Add parameters for each tank                                                # DONE
#[TODO] Add GUI to input parameters for each tank                                   # DONE
#[TODO] Add a Note row for WindowInput Control (DMSO x%)                            # DONE
#[TODO] Automatically save parameters each time the task is changed                 # DONE
#[TODO] Analyzer button will analyze the task being selected                        # DONE
#[TODO] The Cancel button currently create the project at cwd()                     # DONE
#[TODO] Let user select the range of frames within data to be analyzed              #
#[TODO] The number of the nested parameters should be dynamic                       # DONE
#[TODO] The mirror position of the tanks -> need to change the comparing conditions # DONE
#[TODO] Add a (left/right) option to RIGHT side of nested parameters                # DONE
#[TODO] Add a small button to LEFT side of nested parameters to delete              # DONE
#[TODO] Add a batch selector                                                        # DONE
# [BUG] A bug when just created a project then create another one then cancel       # FIXED
# [BUG] Adding more than 2 treatments causing display bug                           # FIXED
#[TODO] Change hyp format to only "MIRROR" & "SEPARATOR", " ZONE" is calculated     # DONE
#[TODO] Note can be edit directly by user, even after the Project creating step     #
#[TODO] Save before analyze                                                         # DONE
#[TODO] Add batch and treatment name to the excel file                              # DONE
#[TODO] Hyperparamater sets for Batch/TreatmentGroup                                # DONE
#[TODO] Import video feature                                                        # DONE
#[TODO] Put units for each parameters                                               # DONE
#[TODO] Change the dropdown menu of ppm/ppb to manual input                         # 
#[TODO] Batch menu moved next Loaded Project, ADD / DELETE buttons                  # DONE
#[TODO] Change to InDetail checkbox to a button that copy the current treatment     # DONE
#[TODO] Import Project from other directory                                         # 
#[TODO] In Tests that has 2 nested params, just 1 add&remove pair of button is ok   # DONE
#[TODO] CREATE PROJECT AUTO LOAD                                                    # DONE
#[TODO] CHANGE PROJECT AUTO SET DEFAULT< IF CAN"T< SEARCH FOR THEM                  # DONE
#[TODO] ALERT BEFORE DELETE PROJECT                                                 # DONE  
#[TODO] Add button don't remove directory if the current number is > target number  # DONE
#[TODO] Parameters used also summarized in a sheet named "Analysis Info"            # 
#[TODO] Change the word "Separator" to something more informative                   # DONE
#[TODO] When the number of fish in sheets are different, there would be issue !     # DONE
#[TODO] Add limitation for misc.clean_df so it does not remove all NaN              # DONE
#[TODO] For NovelTankTest, fill method should have 2 options: "ffill" and "bfill"   # DONE
# -> split the 21000 into 7 x 3000 first, then clean_df
#[TODO] Right before analyzing, check if all trajectories are available, if empty,
# then ask user if the fish there just not moving? If Yes -> ask User to input a X,Y
# If No -> put a blank line in the excel file.                                      # DONE
#[TODO] Bug fix of Control group also has 11 fishes while input only 10?            # 
#[TODO] Remove the import trajectories from the main program, not a live feature    #
#[TODO] Treatment folder names missing substance name, only concentration appears   #
#[TODO] Add the Measurer to calculate the parameters                                #
#[TODO] Make a stop thread checkpoint and a window to ask user 
# if they want to keep the data analyzed partially or not                           #

# For the loader.py purpose
with open('loaded.txt', 'w') as f:
    f.write('loaded')


ROOT = Path(__file__).parent
# ORI_HYP_PATH = ROOT / "Bin"
HISTORY_PATH = "History/projects.json"
init_history(HISTORY_PATH)




# SETUP LOGGING CONFIGURATION
logger = logging.getLogger(__name__)

# save log to Log/log.txt
Path('Log').mkdir(parents=True, exist_ok=True)

# Configure the logging module
log_file = 'Log/log.txt'

class ContextFilter(logging.Filter):
    """
    This is a filter which injects contextual information into the log.
    """

    def filter(self, record):
        record.pathname = os.path.basename(record.pathname)  # Modify this line if you want to alter the path
        return True

# Define the log format with colors
log_format = "%(asctime)s %(log_color)s%(levelname)-8s%(reset)s [%(pathname)s] %(message)s"

# Create a formatter with colored output
formatter = ColoredFormatter(log_format)

# Get the root logger
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

# Create a filter
f = ContextFilter()

# Create a file handler to save logs to the file
file_handler = logging.FileHandler(log_file, mode='a')  # Set the mode to 'a' for append
file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)-8s [%(pathname)s] %(message)s"))
file_handler.addFilter(f)  # Add the filter to the file handler
file_handler.setLevel(logging.DEBUG)

# Create a stream handler to display logs on the console with colored output
stream_handler = logging.StreamHandler()
stream_handler.setFormatter(formatter)
stream_handler.addFilter(f)  # Add the filter to the stream handler
stream_handler.setLevel(logging.DEBUG)

# Add the handlers to the logger
logger.addHandler(file_handler)
logger.addHandler(stream_handler)

class HISTORY():

    def __init__(self, history_path = HISTORY_PATH):
        self.history_path = history_path

        with open(HISTORY_PATH, "r") as file:
            self.projects_data = json.load(file)


    def reload(self):
        with open(HISTORY_PATH, "r") as file:
            self.projects_data = json.load(file)


    def get_project_dir(self, project_name):
        self.reload()

        if project_name == "":
            logger.warning("Tried to get project directory of an empty project name")
            return None
        
        project_dir = self.projects_data[project_name]["DIRECTORY"]

        # check if the project directory exists
        if not os.path.exists(project_dir):
            tkinter.messagebox.showerror("Error", "Project directory does not exist!")
            logger.info(f"Project directory of {project_name} does not exist. Asking for relocation")
            relocate = tkinter.messagebox.askyesno("Project not found", "Do you want to relocate it?")
            if relocate:
                # ask for new input of project_dir
                new_dir = tkinter.filedialog.askdirectory()
                
                self.projects_data[project_name]["DIRECTORY"] = new_dir
                self.saver()
                logger.info(f"Project directory of {project_name} has been relocated to {new_dir}")

                return new_dir
            else:
                return None

        return project_dir
    
    def get_treatment_dir(self, project_name, test_name, batch_num, treatment_char):
        project_path = Path(self.get_project_dir(project_name))

        test_path = [child for child in project_path.iterdir() if child.is_dir() and test_name.lower() in child.name.lower()][0]

        batch_ord = f"{ORDINALS[int(batch_num)-1]} Batch"

        treatment_char = f"{treatment_char} - "
        treatment_path = [child for child in test_path.iterdir() if child.is_dir() and batch_ord in child.name and treatment_char in child.name][0]

        return treatment_path


    def update_blank_folders(self, project_name, test_num, batch_num, treatment_num, target_amount, task):
        project_path = Path(self.get_project_dir(project_name))

        # find children directory of project_path
        test_dirs = [child for child in project_path.iterdir() if child.is_dir() and child.name != "static"]

        batch_ordinal = ORDINALS[batch_num - 1]

        test_name = TESTS_LIST[test_num]
        logger.debug("Test name", test_name)

        for test_dir in test_dirs:
            if test_name not in test_dir.name:
                continue
            # treatment_dir pattern = "A - Control (1st Batch)"
            pattern = f"*{treatment_num}*({batch_ordinal} Batch)"
            logger.debug("Pattern: ", pattern)
            treatment_dirs = test_dir.glob(pattern)
            treatment_dir = list(treatment_dirs)[0]

            # find number of directories inside treatment_dir
            current_fish_list = [child for child in treatment_dir.iterdir() if child.is_dir()]
            # WindowsPath('D:/TestSave/DifNum/04 - Social Interaction Test/A - Control (1st Batch)/1')
            
            current_fish_list = [int(str(fish.name).split("_")[0].strip()) for fish in current_fish_list]
            logger.debug("Current fish list: ", current_fish_list)
            # [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
            
            max_fish_num = max(current_fish_list)
            # 10
            logger.debug("Current fish number: ", max_fish_num)
            logger.debug("Target amount: ", target_amount)
            if max_fish_num < target_amount:
                logger.debug("Current number of fish is less than target amount")
                for i in range(max_fish_num + 1, target_amount + 1):
                    new_fish_dir = treatment_dir / f"{i}"
                    logger.debug("Try creating new fish directory: ", new_fish_dir)
                    new_fish_dir.mkdir()
                    logger.debug(f"New fish directory '{new_fish_dir}' in {test_dir}/{treatment_dir}")
            elif max_fish_num > target_amount:
                logger.debug("Current number of fish is more than target amount")
                if task == "add":
                    logger.debug("Task is 'add', so do nothing")
                    pass
                elif task == "remove":
                    logger.debug("Task is 'remove', so remove fish directories")
                    for i in range(max_fish_num, target_amount, -1):
                        fish_dir = treatment_dir / f"{i}"
                        logger.debug("Try removing fish directory: ", fish_dir)
                        shutil.rmtree(fish_dir)
                        logger.debug(f"Fish directory '{fish_dir}' has been removed")

    def fish_adder(self, project_name, test_num, batch_num, treatment_num, target_amount, task, modify_history=False):
        if treatment_num == "all":
            # get list of treatments in batch_num
            project_detail = self.projects_data[project_name]
            batch_name = f"Batch {batch_num}"
            treatments_count = len(project_detail[batch_name].keys())
            treatment_nums = [chr(i) for i in range(65, 65+treatments_count)]
            for treatment_num in treatment_nums:
                self.add_fish(project_name, test_num, batch_num, treatment_num, target_amount, task, modify_history)
        else:
            self.add_fish(project_name, test_num, batch_num, treatment_num, target_amount, task, modify_history)


    def add_fish(self, project_name, test_num, batch_num, treatment_num, target_amount, task, modify_history=False):
        logger.debug("add_fish is called")
        logger.debug(f"variables: project_name = {project_name}, test_num = {test_num}, batch_num = {batch_num}, treatment_num = {treatment_num}, target_amount = {target_amount}, modify_history = {modify_history}")
        project_detail = self.projects_data[project_name]

        batch_name = f"Batch {batch_num}"
        treatment_name = f"Treatment {treatment_num}"

        if batch_name not in project_detail.keys():
            ERROR = f"Batch {batch_num} does not exist in project {project_name}"
            logger.warning(ERROR)
            return ERROR
        
        if treatment_name not in project_detail[batch_name].keys():
            ERROR = f"Treatment {treatment_num} does not exist in batch {batch_num}"
            logger.warning(ERROR)
            return ERROR
        
        # treatment_detail pattern
        # [
        #         "Melamine",
        #         10.0,
        #         "ppm",
        #         10, # fish num
        #         ""
        #     ]
        # Position of fish num in treatment detail
        DATA_NO = 3

        if DATA_NO > len(project_detail[batch_name][treatment_name]):
            ERROR = f"{DATA_NO} out-of-range for {treatment_num}"
            logger.warning(ERROR)
            return ERROR
        
        target_amount = int(target_amount)

        if modify_history:
            self.projects_data[project_name][batch_name][treatment_name][DATA_NO] = target_amount
            self.saver()

        # Add blank folder to project directory
        self.update_blank_folders(project_name, test_num, batch_num, treatment_num, target_amount, task)

    def saver(self):
        with open(HISTORY_PATH, "w") as file:
            json.dump(self.projects_data, file, indent=4)

THE_HISTORY = HISTORY()

class CustomDialog(tkinter.Toplevel):
    def __init__(self, master, title=None, message=None, button_text=None, button_command=None):
        tkinter.Toplevel.__init__(self, master)
        self.title(title)

        self.label = tkinter.Label(self, text=message)
        self.label.pack(padx=10, pady=10)

        self.button_command = button_command
        self.button = tkinter.Button(self, text=button_text, command=self.ok)
        self.button.pack(pady=10)

        self.geometry("+%d+%d" % (master.winfo_rootx(), master.winfo_rooty()))

    def ok(self):
        if self.button_command is not None:
            self.button_command()
        self.destroy()


class ToolTip(object):

    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 57
        y = y + cy + self.widget.winfo_rooty() +27
        self.tipwindow = tw = tkinter.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tkinter.Label(tw, text=self.text, justify=tkinter.LEFT,
                      background="#ffffe0", relief=tkinter.SOLID, borderwidth=1,
                      font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

def CreateToolTip(widget, text):
    toolTip = ToolTip(widget)
    def enter(event):
        toolTip.showtip(text)
    def leave(event):
        toolTip.hidetip()
    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)


class ScrollableProjectList(customtkinter.CTkScrollableFrame):

    def __init__(self, master, command=None, **kwargs):

        super().__init__(master, **kwargs)
        self.grid_columnconfigure(0, weight=1)

        self.command = command
        self.project_variable = customtkinter.StringVar()
        self.project_radiobuttons = []

    def add_project(self, project_name):
        project_radiobutton = customtkinter.CTkRadioButton(
            self, text=project_name, value=project_name, variable=self.project_variable
        )
        project_radiobutton.grid(row=len(self.project_radiobuttons), column=0, pady=(0, 10), sticky="w")
        self.project_radiobuttons.append(project_radiobutton)

    def clear_projects(self):
        for radiobutton in self.project_radiobuttons:
            radiobutton.destroy()
        self.project_radiobuttons = []

    def get_selected_project(self):
        return self.project_variable.get()

    def set_selected_project(self, project_name="last"):
        if project_name == "last":
            # set to the last project in list
            self.project_variable.set(self.project_radiobuttons[-1].cget("text"))
            logger.warning("Set project variable failed, set to the last project in list")
        else:
            self.project_variable.set(project_name)
            logger.debug("Set project variable to " + project_name)
    
    def select_project(self, project_name):
        for radiobutton in self.project_radiobuttons:
            if radiobutton.cget("text") == project_name:
                radiobutton.invoke()
                break

    def return_recent_project(self):
        return self.project_radiobuttons[-1].cget("text")
        


class ProjectDetailFrame(customtkinter.CTkFrame):

    def __init__(self, master, project_name, **kwargs):

        super().__init__(master, **kwargs)

        # # Create tree view
        # self.tree = ttk.Treeview(self, height = 5, show = "headings")
        # self.tree.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        

        # If project name is not empty, load the project details, 
        # otherwise, display "No project selected"
        self.project_name = project_name
        if self.project_name != "":
            self.load_project_details()
        else:
            label = customtkinter.CTkLabel(self, text="No project selected")
            label.grid(row=0, column=0, padx=5, pady=5)

    def update_grid_weight(self):
        rows, cols = self.grid_size()

        for row in range(rows):
            self.grid_rowconfigure(row, weight=1)

        for col in range(cols):
            self.grid_columnconfigure(col, weight=1)

    def load_project_details(self, project_name=None, batch_name="Batch 1"):

        logger.info(f"Loading.. project name = {project_name}")

        if project_name == "":
            label = customtkinter.CTkLabel(self, text="No project selected")
            label.grid(row=0, column=0, padx=5, pady=5)
            return

        if project_name is not None:
            self.project_name = project_name

        with open("History/projects.json", "r") as file:
            projects_data = json.load(file)

        project_data = projects_data[self.project_name][batch_name]

        logger.info(project_data)

        headers = ["Treatment", "Dose", "Dose Unit", "Fish Number", "Note"]

        # scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        # scroll.grid(row=0, column=1, sticky="ns")  # Changed from scroll.pack to scroll.grid

        # self.tree.configure(yscrollcommand=scroll.set)

        # for i, header in enumerate(headers):
        #     self.tree.heading(i, text=header)
        #     self.tree.column(i, width=100, anchor='center')

        # for details in project_data.values():
        #     treatment_name, dose, dose_unit, fish_number, note = details

        #     dose = dose if dose != 0 else ""
        #     dose_unit = dose_unit if dose_unit != "" else ""
        #     fish_number = fish_number if fish_number != 0 else ""

        #     labels = [treatment_name, dose, dose_unit, fish_number, note]

        #     self.tree.insert("", "end", values=labels)

        for i, header in enumerate(headers):
            label = customtkinter.CTkLabel(self, text=header, font=customtkinter.CTkFont(weight="bold"))
            label.grid(row=0, column=i, padx=5, pady=5)

        for row, (treatment, details) in enumerate(project_data.items(), start=1):
            treatment_name, dose, dose_unit, fish_number, note = details

            dose = dose if dose != 0 else ""
            dose_unit = dose_unit if dose_unit != "" else ""
            fish_number = fish_number if fish_number != 0 else ""

            labels = [treatment_name, dose, dose_unit, fish_number, note]

            for col, label_text in enumerate(labels):
                label = customtkinter.CTkLabel(self, text=label_text)
                label.grid(row=row, column=col, padx=5, pady=5)

        self.update_grid_weight()

    def clear(self):
        for child in self.winfo_children():
            child.destroy()


class Parameters(customtkinter.CTkFrame):

    def __init__(self, master, project_name, selected_task, nested_key, *args, **kwargs):
        
        super().__init__(master, *args, **kwargs)

        self.project_name = project_name
        self.selected_task = selected_task

        self.null_label = None

        if self.project_name == "":
            self.null_label = customtkinter.CTkLabel(self, text="No project selected")
            self.null_label.grid(row=0, column=0, padx=5, pady=5)
        else:
            self.load_parameters(nested_key = nested_key)

        self.entries = {}

        self.UNITS = {
            "FRAME RATE": "frames/second",
            "DURATION": "seconds",
            "CONVERSION RATE": "pixels/cm",
            "ZONE WIDTH": "cm",
            "SEGMENT DURATION": "seconds",
        }

    def get_hyp_path(self, project_dir, selected_task_init, condition, batch_num, treatment_mode = 'current'):

        assert treatment_mode in ['all', 'current']

        hyp_name = f"hyp_{selected_task_init}.json"

        if treatment_mode == 'current':

            hyp_path = os.path.join(project_dir, 'static', f"Batch {batch_num}", condition, hyp_name)

            logger.debug(f"Retrieved hyp from {hyp_path}, mode = {treatment_mode}")

            return [hyp_path]
        
        elif treatment_mode == 'all':
            
            hyp_paths = []

            hyp_batch_dir = Path(os.path.join(project_dir, 'static', f"Batch {batch_num}"))

            # find all subdirectory in hyp_batch_dir
            for sub_dir in hyp_batch_dir.iterdir():
                if sub_dir.is_dir():
                    hyp_path = sub_dir / hyp_name
                    hyp_paths.append(hyp_path)

            logger.debug(f"Retrieved hyps from {hyp_paths}, mode = {treatment_mode}")

            return hyp_paths

    def get_current_entry_quantity(self):

        last_row = list(self.entries.keys())[-1]
        last_entry = self.entries[last_row]

        try:
            last_row_num = int(last_row.split('_')[-1])
            return last_row_num
        except:
            return 0
        
    def add_entry(self):

        # add a new entry to self.entries, similar to the last entry
        # add a new row to the grid
        last_row = list(self.entries.keys())[-1]
        last_entry = self.entries[last_row]

        try:
            last_row_num = int(last_row.split('_')[-1])
        except:
            return 0

        new_row_num = last_row_num + 1
        new_row = f"{last_row.split('_')[0]}_{new_row_num}"

        try:
            entry_length = len(last_entry)
        except:
            entry_length = 1

        if entry_length == 1:
            value_entry = customtkinter.CTkEntry(self)
            value_entry.insert(0, 0)
            value_entry.grid(row=new_row_num, column=1, padx=5, pady=5)
            
            new_entry = value_entry

        elif entry_length == 2:
            value_entry = customtkinter.CTkEntry(self)
            value_entry.insert(0, 0)
            value_entry.grid(row=new_row_num, column=1, padx=5, pady=5)

            LR_switch = customtkinter.CTkSwitch(self, text=None)
            LR_switch.grid(row=new_row_num, column=2, padx=(5,0), pady=5)

            new_entry = [value_entry, LR_switch]

        self.key_labels[str(new_row_num)] = customtkinter.CTkLabel(self, text=str(new_row_num))
        self.key_labels[str(new_row_num)].grid(row=new_row_num, column=0, padx=5, pady=5)
        self.entries[new_row] = new_entry
        logger.info(f"Added entry {new_row}")

        return new_row_num


    def remove_entry(self):
        last_row = list(self.entries.keys())[-1]
        last_entry = self.entries[last_row]

        try:
            last_row_num = int(last_row.split('_')[-1])
        except:
            return
        
        try:
            entry_length = len(last_entry)
        except:
            entry_length = 1

        if entry_length == 1:
            last_entry.destroy()
        
        elif entry_length == 2:
            last_entry[0].destroy()
            last_entry[1].destroy()

        # remove the last entry from self.entries
        self.key_labels[str(last_row_num)].destroy()
        self.key_labels.pop(str(last_row_num))
        self.entries.pop(last_row)
        logger.info(f"Removed entry {last_row}")

        return last_row_num-1
        

    def load_parameters(self, project_name=None, selected_task=None, condition = None, batch_num=None, nested_key=0):
        logger.debug(f"Loading parameters for project_name = {project_name}, selected_task = {selected_task}, condition = {condition}, batch_num = {batch_num}, nested_key = {nested_key}")

        self.null_label = None

        self.entries = {}

        self.clear()

        if project_name == None:
            project_name = self.project_name

        if selected_task is not None:
            self.selected_task = selected_task

        self.selected_task = self.selected_task.split()[0].lower()
        hyp_name = f"hyp_{self.selected_task}.json"

        if project_name == "":
            # self.hyp_path = ORI_HYP_PATH / hyp_name
            ori_dict = CONSTANTS[hyp_name]
        else:
            project_dir = THE_HISTORY.get_project_dir(project_name)
            if project_dir == None:
                logger.warning(f"Project {project_name} not found and user don't know where to find it")
                return None
            self.hyp_path = self.get_hyp_path(project_dir, self.selected_task, condition, batch_num)[0]
            with open(self.hyp_path, "r") as file:
                ori_dict = json.load(file)

        # find the nested keys
        nested_keys = []
        for key, value in ori_dict.items():
            if isinstance(value, dict):
                nested_keys.append(key)

        if nested_key == 0:
            display_dict = {k: v for k, v in ori_dict.items() if not isinstance(v, (dict, list))}
            headers = ["Parameter", "Value", "Unit"]
        else:
            try:
                nested_key = nested_keys[nested_key-1]
                

            except IndexError:
                self.null_label = customtkinter.CTkLabel(self, text="No more nested keys")
                self.null_label.grid(row=0, column=0, padx=5, pady=5) 
                return "None"
            
            display_dict = ori_dict[nested_key]

            example_value = list(display_dict.values())[0]
            if isinstance(example_value, list):
                headers = ["Tank", "Value", "Left/Right"]
            else:
                headers = ["Tank", "Value"]

        example_key = list(display_dict.keys())[0]
        try:
            _ = int(example_key)
            pass
        except ValueError:
            units = [self.UNITS[k] for k in display_dict.keys()]
            for i, unit in enumerate(units):
                unit_label = customtkinter.CTkLabel(self, text=unit)
                unit_label.grid(row=i+1, column=2, padx=(5,10), pady=5)

        self.key_labels = {}

        for row, (key, value) in enumerate(display_dict.items()):
            self.key_labels[key] = customtkinter.CTkLabel(self, text=key)
            self.key_labels[key].grid(row=row+1, column=0, padx=5, pady=5)

            # if value is a list
            if isinstance(value, list):
                LR_switch = customtkinter.CTkSwitch(self, text=None)
                LR_switch.grid(row=row+1, column=2, padx=(5,0), pady=5)
                if int(value[1]) == 1:
                    LR_switch.select()

                display_value = value[0]
                # headers = ["Tank", "Value", "Left/Right"]
            else:
                display_value = value
                # headers = ["", "Value"]

            value_entry = customtkinter.CTkEntry(self)
            value_entry.insert(0, display_value)
            value_entry.grid(row=row+1, column=1, padx=5, pady=5)

            try:
                _ = int(key)
                entry_key = f"{nested_key}_{key}"
            except ValueError:
                entry_key = key

            if isinstance(value, list):
                self.entries[entry_key] = [value_entry, LR_switch]
            else:
                self.entries[entry_key] = value_entry

        # make a header
            for i, header in enumerate(headers):
                label = customtkinter.CTkLabel(self, text=header, font=customtkinter.CTkFont(weight="bold"))
                label.grid(row=0, column=i, padx=5, pady=5)


        return nested_key

    def clear(self):
        for child in self.winfo_children():
            child.destroy()

    def save_parameters(self, project_name, selected_task, condition, batch_num, treatment_mode = 'all'):
        logger.debug(f"Saving parameters for {project_name}.{selected_task}.{condition}.{batch_num} in mode = {treatment_mode}")

        assert treatment_mode in ['all', 'current'] 

        def get_entry(entry_dict):
            out_dict = {}
            for key, value in entry_dict.items():
                try:
                    if isinstance(value, list):
                        v = [value[0].get(), value[1].get()]
                    else:
                        v = value.get()
                except AttributeError:
                    logger.warning(f"During saving parameters for {project_name}.{selected_task}.{condition}.{batch_num} in mode = {treatment_mode}")
                    logger.warning(f"AttributeError: {key} is not a tkinter entry")
                    logger.warning(f"Value: ", v)
                    logger.warning(f"Value type: ", type(v))
                    continue
                out_dict[key] = v
            return out_dict
        
        selected_task = self.selected_task.split()[0].lower()

        if project_name == "":
            tkinter.messagebox.showerror("Warning", "No project selected. No save was made.")
            return            
        else:
            project_dir = THE_HISTORY.get_project_dir(project_name)
            hyp_paths = self.get_hyp_path(project_dir, self.selected_task, condition, batch_num, treatment_mode=treatment_mode)


        # Get the values from the entries
        updated_values = get_entry(self.entries)
        
        for hyp_path in hyp_paths:

            # load the original data
            with open(hyp_path, "r") as file:
                parameters_data = json.load(file)

            #separate the updated_values.items into 2 groups
            updated_values_simple = {key: value for key, value in updated_values.items() if "_" not in key}
            updated_values_nested = {key: value for key, value in updated_values.items() if "_" in key}
            updated_values_nested_grouped = {}
            for key, value in updated_values_nested.items():
                nested_key, nested_key_fish = key.split("_")
                if nested_key not in updated_values_nested_grouped:
                    updated_values_nested_grouped[nested_key] = {}
                updated_values_nested_grouped[nested_key][nested_key_fish] = value

            # Update the values in the dictionary with the new values
            for key, value in updated_values_simple.items():
                try:
                    parameters_data[key] = value
                except ValueError:
                    logger.warning(f"Invalid input for {key}: {value}. Skipping.")
            
            for nested_key, fishes in updated_values_nested_grouped.items():
                if nested_key not in parameters_data:
                    logger.error(f"Nested key {nested_key} not found in {hyp_path}")
                    raise ValueError(f"Nested key {nested_key} not found in {hyp_path}")
                parameters_data[nested_key] = fishes


            # # Update the values in the dictionary with the new values
            # for key, value in updated_values.items():
            #     try:
            #         if "_" not in key:
            #             parameters_data[key] = value
            #         else:
            #             nested_key, nested_key_fish = key.split("_")
            #             if nested_key not in parameters_data:
            #                 parameters_data[nested_key] = {}
            #             parameters_data[nested_key][nested_key_fish] = value

            #     except ValueError:
            #         print(f"Invalid input for {key}: {value}. Skipping.")

            # Save the updated data to the file
            with open(hyp_path, "w") as file:
                json.dump(parameters_data, file, indent=4)

            logger.info(f"Parameters of {selected_task} saved to {hyp_path}.")

class NK_button(customtkinter.CTkButton):
    def __init__(self, parent, text, command, row, column, columnspan=1, *args, **kwargs):
        super().__init__(parent, text=text, command=command, *args, **kwargs)
        self.parent = parent
        self.text = text
        self.command = command
        self.row = row
        self.column = column
        self.columnspan = columnspan

    def show(self):
        self.grid(row=self.row, 
                  column=self.column, 
                  columnspan=self.columnspan, 
                  padx=5, 
                  pady=5, 
                  sticky="nsew")

    def hide(self):
        self.grid_forget()


class App(customtkinter.CTk):

    def __init__(self):

        super().__init__()

        # PREDEFINED VARIABLES
        self.PROJECT_CREATED = False
        self.CURRENT_PROJECT = ""
        # self.CURRENT_PARAMS = {}
        self.TESTLIST = TESTS_LIST
        self.PREVIOUS_TEST = ""
        self.PREVIOUS_BATCH = ""
        self.PREVIOUS_CONDITION = ""
        self.PREVIOUS_DIFFERENCE = 0
        self.CONDITIONLIST = ["Treatment A", "Treatment B", "Treatment C"]

        # configure window
        self.title("Tower Assay Analyzer")
        self.geometry(f"{1500}x{790}")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=0) 
        self.grid_columnconfigure((2, 3), weight=1)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        ### COLUMN 0 ###

        button_config = {"font": ('Helvetica', 16), "width": 150, "height": 40}

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="Tower Assay Analyzer", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, columnspan=2, padx=20, pady=(20, 10))
        
        self.sidebar_button_1 = customtkinter.CTkButton(self.sidebar_frame, text="Create Project", 
                                                        command=self.create_project)
        self.sidebar_button_1.configure(**button_config)
        self.sidebar_button_1.grid(row=1, column=0, columnspan=2, padx=20, pady=20)

        self.sidebar_button_2 = customtkinter.CTkButton(self.sidebar_frame, text="Load Project", 
                                                        command=self.load_project)
        self.sidebar_button_2.configure(**button_config)
        self.sidebar_button_2.grid(row=2, column=0, columnspan=2, padx=20, pady=20)

        self.sidebar_button_3 = customtkinter.CTkButton(self.sidebar_frame, text="Delete Project", 
                                                        command=self.delete_project)
        self.sidebar_button_3.configure(**button_config)
        self.sidebar_button_3.grid(row=3, column=0, columnspan=2, padx=20, pady=20)

        # self.batch_label = customtkinter.CTkLabel(self.sidebar_frame, text="Batch Number", font=customtkinter.CTkFont(size=16))
        # self.batch_label.grid(row=4, column=0, padx=5, pady=5)
        # self.batch_entry = customtkinter.CTkEntry(self.sidebar_frame, width=50, height=10)
        # self.batch_entry.grid(row=4, column=1, padx=5, pady=5)
        # # set default value = 1
        # self.batch_entry.insert(0, "1")

        self.ImportVideoButton = customtkinter.CTkButton(self.sidebar_frame, text="Import Video",
                                                   command=self.import_video)
        self.ImportVideoButton.configure(**button_config)
        self.ImportVideoButton.grid(row=4, column=0, columnspan = 2, padx=20, pady=20)

        self.sidebar_button_4 = customtkinter.CTkButton(self.sidebar_frame, text="Import Trajectories", 
                                                        command=self.import_trajectories)
        self.sidebar_button_4.configure(**button_config)
        self.sidebar_button_4.grid(row=5, column=0, columnspan=2, padx=20, pady=20)

        self.sidebar_button_5 = customtkinter.CTkButton(self.sidebar_frame, text="Analyze", 
                                                        command=self.analyze_project_THREADED)
        self.sidebar_button_5.configure(**button_config)
        self.sidebar_button_5.grid(row=6, column=0, columnspan=2, padx=20, pady=20)


        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=7, column=0, columnspan=2, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        
        self.appearance_mode_optionemenu.grid(row=8, column=0, columnspan=2, padx=20, pady=(10, 10))
        
        self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        self.scaling_label.grid(row=9, column=0, columnspan=2, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
                                                               command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=10, column=0, columnspan=2, padx=20, pady=(10, 20))

        ### COLUMN 1 ###

        container_1 = customtkinter.CTkFrame(self)
        container_1.grid(row=0, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")

        container_1.grid_rowconfigure(0, weight=0)
        container_1.grid_rowconfigure(1, weight=1)
        container_1.grid_columnconfigure(0, weight=0)

        # Top part
        container_2_top = customtkinter.CTkFrame(container_1)
        container_2_top.grid(row=0, column=0, sticky="nsew")

        project_previews_label = customtkinter.CTkLabel(container_2_top, text="Project List", font=customtkinter.CTkFont(size=20, weight="bold"))
        project_previews_label.grid(row=0, column=0)

        # Bottom part
        bottom_part = customtkinter.CTkFrame(container_1)
        bottom_part.grid(row=1, column=0, sticky="nsew")

        bottom_part.grid_rowconfigure(0, weight=1)
        bottom_part.grid_rowconfigure(1, weight=0)

        self.scrollable_frame = ScrollableProjectList(bottom_part)
        self.scrollable_frame.grid(row=0, column=0, sticky="nsew")

        refresh_button = customtkinter.CTkButton(bottom_part, text="Refresh", command=self.refresh_projects)
        refresh_button.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")

        # Initial refresh to populate the list
        self.refresh_projects()

        self.project_detail_container = ProjectDetailFrame(self, self.CURRENT_PROJECT, width = 400)
        self.project_detail_container.grid(row=1, column = 1, columnspan=3, padx=20, pady=20, sticky="nsew")

        ### COLUMN 2 ###

        # Create a canvas to hold the project parameters
        container_2 = customtkinter.CTkFrame(self, width = 400)
        container_2.grid(row=0, column=2, columnspan = 2, padx=(20, 0), pady=(20, 0), sticky="nsew")

        # ROW 0
        # Top part is a dropdown menu to select type of test
        container_2_top = customtkinter.CTkFrame(container_2)
        container_2_top.grid(row=0, column=0, columnspan=3, sticky="nsew")

        Header = customtkinter.CTkLabel(container_2_top, text="Loaded Project:", anchor="w", font=customtkinter.CTkFont(size=15, weight="bold"))
        Header.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="nsew")
        self.LoadedProject = customtkinter.CTkLabel(container_2_top, text="None", anchor="w", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.LoadedProject.grid(row=0, column=1, columnspan=2, padx=20, pady=(20, 10), sticky="nsew")

        # ROW 1
        self.BATCHLIST = ["Batch 1"]

        self.container_2_mid = customtkinter.CTkFrame(container_2)
        self.container_2_mid.grid(row=1, column=0, columnspan=3, sticky="nsew")

        
        self.BatchOptions = customtkinter.CTkOptionMenu(self.container_2_mid, dynamic_resizing=False,
                                                        width = 105, values=self.BATCHLIST)
        self.BatchOptions.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="nsew")

        self.BatchAddButton = customtkinter.CTkButton(self.container_2_mid, text="Add Batch", width = 40,
                                                        command=self.add_batch)
        self.BatchAddButton.grid(row=0, column=1, padx=20, pady=(20, 10), sticky="nsew")

        self.BatchRemoveButton = customtkinter.CTkButton(self.container_2_mid, text="Remove Batch", width = 40,
                                                        command=self.remove_batch)
        self.BatchRemoveButton.grid(row=0, column=2, padx=20, pady=(20, 10), sticky="nsew")
        
        self.TestOptions = customtkinter.CTkOptionMenu(self.container_2_mid, dynamic_resizing=False, 
                                                  width=210, values=self.TESTLIST)
        self.TestOptions.grid(row=1, column=0, columnspan = 2, padx=20, pady=(20, 10), sticky="nsew")

        self.save_button = customtkinter.CTkButton(self.container_2_mid, text="Save", width = 50,
                                                   command=self.save_parameters)
        self.save_button.grid(row=1, column=2, padx=20, pady=20, sticky="nsew")

        self.TreatmentOptions = customtkinter.CTkOptionMenu(self.container_2_mid, dynamic_resizing=False,
                                                                width=210, values=self.CONDITIONLIST)
        self.TreatmentOptions.grid(row=2, column=0, columnspan=3, padx=20, pady=(20, 10), sticky="nsew")

        self.parameters_frame = Parameters(self.container_2_mid, self.CURRENT_PROJECT, self.TESTLIST[0], 0)
        self.parameters_frame.grid(row=3, columnspan=3, padx=20, pady=20, sticky="nsew")

        # Row 2
        container_2_bot = customtkinter.CTkFrame(container_2)
        container_2_bot.grid(row=2, column=0, columnspan=3, sticky="nsew")

        # create a Cloner button to copy current treatment's parameters to other treatment
        self.Cloner = customtkinter.CTkButton(container_2_bot, text="Copy to other treatment", width = 50,
                                              command=self.copy_to_other_treatment)
        self.Cloner.grid(row=1, column=0, pady=20, padx=20, sticky="nsew")

        self.ClonerToolTip = tkinter.Button(container_2_bot, text="?")
        self.ClonerToolTip.grid(row=1, column=1, pady=20, padx=20)
        CreateToolTip(self.ClonerToolTip, text = 'Copy all parameters setting above and on the right-side columns\n'
                 'to other Treatment and save them immediately\n'
        )

        self.CheckIntegrity = customtkinter.CTkButton(container_2_bot, text="Trajectories Check", width = 50,
                                                command=self.trajectories_check)
        self.CheckIntegrity.grid(row=1, column=2, pady=20, padx=20)


        ### COLUMN 3+ ###

        # ROW 0 #

        container_3 = customtkinter.CTkScrollableFrame(self, width = 500)
        container_3.grid(row=0, rowspan=2, column=5, columnspan = 2, padx=(20, 0), pady=(20, 0), sticky="nsew")

        self.nested_key_1_header = customtkinter.CTkLabel(container_3, text="None", anchor="w", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.nested_key_1_header.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="nsew")
        self.nested_key_1_frame = Parameters(container_3, self.CURRENT_PROJECT, self.TESTLIST[0], 1)
        self.nested_key_1_frame.grid(row=1, column=0, columnspan=2, padx=20, pady=(10, 20), sticky="nsew")

        self.nk_add_button = NK_button(container_3, text="Add", width = 20,
                                        row = 2, column = 0,
                                        command=self.nk_add)
        self.nk_remove_button = NK_button(container_3, text="Remove", width = 20,
                                        row = 2, column = 1,
                                        command=self.nk_remove)
        # self.nk_import_button = NK_button(container_3, text="Import", width=20,
        #                                 row=3, column=0, 
        #                                 command=self.import_element_dialog)
        self.nk_selector_button = NK_button(container_3, text="Select from image", width=20,
                                            row=3, column=0, columnspan=2,
                                            command=self.nk_selector_dialog)

        self.nested_key_2_header = customtkinter.CTkLabel(container_3, text="None", anchor="w", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.nested_key_2_header.grid(row=0, column=2, padx=20, pady=(20, 10), sticky="nsew")
        self.nested_key_2_frame = Parameters(container_3, self.CURRENT_PROJECT, self.TESTLIST[0], 2)
        self.nested_key_2_frame.grid(row=1, column=2, columnspan = 2, padx=20, pady=(10, 20), sticky="nsew")

        # Config
        self.BatchOptions.configure(command=self.update_param_display)
        self.TestOptions.configure(command=self.update_param_display)
        self.TreatmentOptions.configure(command=self.update_param_display)

        # Load the first test by default
        self.update_param_display(load_type = "first_load")

        # logger.critical(f"Entry frame in middle: {self.parameters_frame.entries}")
    


    def get_treatment_char(self, current_condition = None):
        if current_condition == None:
            current_condition = self.TreatmentOptions.get()
        condition_index = self.CONDITIONLIST.index(current_condition)
        condition_char = CHARS[condition_index]
        return condition_char
    
    def trajectories_check(self, mode = "current"):
        assert mode in ["current", "all"]
        logger.debug(f"Checking the existence of the trajectories, mode = {mode}")

        if self.CURRENT_PROJECT == "":
            tkinter.messagebox.showerror("Error", "No project is currently opened")
            return

        # get current project_dir
        current_project_dir = THE_HISTORY.get_project_dir(self.CURRENT_PROJECT)
        # get current test
        current_test = self.TestOptions.get()
        # get current treatment
        # get current batch
        current_batch = self.BatchOptions.get()
        all_treatment_chars = [self.get_treatment_char(condition) for condition in self.CONDITIONLIST]

        false_keys_dict = {}
        treatments_checker_dict = {}

        if mode == "current":
            current_treatment = self.get_treatment_char()
            all_treatment_chars = [current_treatment]

        logger.debug(f"Checking these treatments: {all_treatment_chars}")

        self.checker_windows = {}

        for i, treatment_char in enumerate(all_treatment_chars):
            checker_dict = check_trajectories_dir(current_project_dir, current_test, treatment_char, current_batch)

            # create a topview window
            self.checker_windows[i] = tkinter.Toplevel(self)
            self.checker_windows[i].title("Trajectories Check")
            self.checker_windows[i].minsize(400, 400)
            self.checker_windows[i].geometry("+%d+%d" % (self.winfo_screenwidth()/2 - 400, self.winfo_screenheight()/2 - 300))
            self.checker_windows[i].rowconfigure(0, weight=1)
            #bring it to front
            self.checker_windows[i].lift()
            self.checker_windows[i].after(5000, self.checker_windows[i].destroy)

            # Create a Canvas widget with a Scrollbar
            checker_container = tkinter.Canvas(self.checker_windows[i])
            checker_container.pack(side=tkinter.LEFT, fill=tkinter.BOTH, expand=True)

            scrollbar = tkinter.Scrollbar(self.checker_windows[i], orient=tkinter.VERTICAL, command=checker_container.yview)
            scrollbar.pack(side=tkinter.RIGHT, fill=tkinter.Y)

            checker_container.configure(yscrollcommand=scrollbar.set)
            checker_container.bind('<Configure>', lambda e: checker_container.configure(scrollregion=checker_container.bbox('all')))

            # Create a frame inside the canvas to hold the labels
            checker_frame = tkinter.Frame(checker_container)
            checker_container.create_window((0,0), window=checker_frame, anchor='nw')

            row = 0
            false_keys = []
            for key, value in checker_dict.items():
                # Create a label for the key
                key_label = tkinter.Label(checker_frame, text=key)
                key_label.grid(row=row, column=0)

                # Create a label for the value
                if value:
                    value_label = tkinter.Label(checker_frame, text="✓")
                else:
                    value_label = tkinter.Label(checker_frame, text="✗")
                    false_keys.append(key)

                value_label.grid(row=row, column=1)

                row += 1

            # Update the scrollregion after creating all the labels
            checker_container.update_idletasks()
            checker_container.configure(scrollregion=checker_container.bbox('all'))

            false_keys_dict[treatment_char] = false_keys
            treatments_checker_dict[treatment_char] = checker_dict

        # From false_keys_dict, remove the keys that has value = []
        false_keys_dict = {key: value for key, value in false_keys_dict.items() if value != []}

        # if all len of values in false_keys_dict == 0, then all true
        if false_keys_dict == {}:
            _status = "All True"
            return _status, None, None
        # if all len of values in false_keys_dict == len(checker_dict), then all false
        elif all([len(value) == len(treatments_checker_dict[key]) for key, value in false_keys_dict.items()]):
            _status = "All False"
            return _status, None, None
        else:
            _status = "Some False"
            return _status, false_keys_dict, treatments_checker_dict


        # if len(false_keys) == 0:
        #     _status = "All True"
        #     return _status, None, None
        # elif len(false_keys) == len(checker_dict):
        #     _status = "All False"
        #     return _status, None, None
        # else:
        #     _status = "Some False"
        #     return _status, false_keys, checker_dict


    def copy_to_other_treatment(self):
        current_treatment = self.TreatmentOptions.get() #OK - just for display

        message_ = f"You are going to copy current treatment parameters to other treatments'"
        message_ += f"\nThis is an irreversible action, do you want to continue?"
        confirm = tkinter.messagebox.askyesno("Confirmation", message_)
        if not confirm:
            return

        logger.debug(f"Copying parameters from {current_treatment} to other treatment")
        treatment_mode = "all"


        # save current treatment parameters to other treatments
        self.save_parameters(mode="current", treatment_mode=treatment_mode)

        # count current entries in available parameters
        target_amount = self.nested_key_1_frame.get_current_entry_quantity()
        # mimic the folder change of the current treatment to other treatment
        self.folder_changer(target_amount, treatment_mode=treatment_mode)

        #Notification
        message_ = "Copied the parameters settings to all other treatments and Saved!"
        tkinter.messagebox("Action Completed", message_)
        logger.debug(message_)


    def nk_selector_dialog(self):
        pass


    def import_element_dialog(self):
        self.import_elem_window = tkinter.Toplevel(self)
        self.import_elem_window.title("Import Elements")

        self.file_path = ""

        def choose_excel(button):
            self.file_path = tkinter.filedialog.askopenfilename(defaultextension=".xlsx", 
                                            filetypes=[("Excel Files", "*.xlsx")])
            logger.debug(f"File path: {self.file_path}")
            # turn button color to black
            button.config(bg="red")
            # turn button text to "imported"
            button.config(text="Excel Loaded")


        # Depending on nested_key_2_frame.null_label, the drop down list will have different options
        options = [self.nested_key_1_header.cget("text")]
        if self.nested_key_2_frame.null_label is None:
            options.append(self.nested_key_2_header.cget("text"))

        chosen_option = tkinter.StringVar(self.import_elem_window)
        chosen_option.set(options[0]) # set the default option

        dropdown = tkinter.OptionMenu(self.import_elem_window, chosen_option, *options)
        dropdown.pack()

        choose_excel_button = tkinter.Button(self.import_elem_window, text="Choose Excel File", command=lambda: choose_excel(choose_excel_button))
        choose_excel_button.pack()

        tkinter.Label(self.import_elem_window, text="Write the exact name of the element you want").pack()

        self.entry_text = tkinter.StringVar()
        import_name_entry = tkinter.Entry(self.import_elem_window, textvariable=self.entry_text)
        import_name_entry.pack()
        

        import_button = tkinter.Button(self.import_elem_window, text="Import", command=lambda: self.import_element(chosen_option.get()))
        import_button.pack()

        def autofill(event=None):
            chosen_text = chosen_option.get()
            # lowercase the chosen_text
            chosen_text = chosen_text.lower()
            # Uppercase the first letter
            chosen_text = chosen_text[0].upper() + chosen_text[1:]
            # insert a default value for import_name_entry = current value of chosen_option
            import_name_entry.insert(0, chosen_text)

        autofill()

        #bind the dropdown menu to autofill the entry
        dropdown.bind("<ButtonRelease-1>", autofill)


    def import_element(self, chosen_option):
        elements = {}
        conversion = 0
        display_message = []
        if self.file_path != "":
            workbook = openpyxl.load_workbook(self.file_path)
            progress_bar = self.create_progress_window(title="Importing Elements")
            for sheet_name in workbook.sheetnames:
                progress = round((workbook.sheetnames.index(sheet_name) + 1) / len(workbook.sheetnames) * 100)
                progress_bar['value'] = progress
                progress_bar.update()
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if str(cell.value).strip() == self.entry_text.get():
                            try:
                                elements[sheet_name] = sheet.cell(row=cell.row+1, column=cell.column).value
                            except IndexError:
                                continue  # Handles the case where the cell is in the last row of the sheet
                        elif str(cell.value).strip() == "Conversion":
                            conversion = sheet.cell(row=cell.row+1, column=cell.column).value

            frame_to_change = self.nested_key_1_frame if chosen_option == self.nested_key_1_header.cget("text") else self.nested_key_2_frame
            difference = len(elements) - len(frame_to_change.entries)
            if difference > 0:
                for _ in range(difference):
                    frame_to_change.add_entry()
            elif difference < 0:
                for _ in range(abs(difference)):
                    frame_to_change.remove_entry()

            logger.debug(f"Elements to import: {elements}") 
            logger.debug(f"Conversion to import: {conversion}")

            for entry, element in zip(frame_to_change.entries.values(), elements.values()):
                try:
                    entry[0].delete(0, tkinter.END)
                    try:
                        entry[0].insert(0, str(element))  # Convert element to string
                    except TypeError as e:
                        logger.error(f"TypeError on insert: {e}")
                        logger.error(f"element is {element}, type(element) is {type(element)}")
                except Exception as e:
                    logger.debug(f"Entry value is {entry[0]}")
                    raise e

                message = f"Imported {element} to {chosen_option}"
                logger.debug(message)
                display_message.append(message)

            # in self.parameters_frame, change the conversion entry
            self.parameters_frame.entries["CONVERSION RATE"].delete(0, tkinter.END)
            self.parameters_frame.entries["CONVERSION RATE"].insert(0, conversion)
            message = f"Imported {conversion} to conversion rate"
            logger.debug(message)
            display_message.append(message)

            progress_bar.master.destroy()

        # Destroy self.import_elem_window
        self.import_elem_window.destroy()

        if len(display_message) > 0:
            tkinter.messagebox.showinfo("Action Completed\n", "\n".join(display_message))


    def folder_changer(self, target_amount, task="add", treatment_mode="current"):
        logger.debug(f"folder_changer called with target_amount = {target_amount}")
        current_test = self.TestOptions.get()
        current_test_index = self.TESTLIST.index(current_test)
        logger.debug(f"current_test_index = {current_test_index}")
        
        current_batch = self.BatchOptions.get()
        current_batch_index = int(current_batch.split(" ")[1])
        logger.debug(f"current_batch_index = {current_batch_index}")

        if treatment_mode == "current":
            current_condition_char = self.get_treatment_char()
            logger.debug(f"Modify folders for current treatment: {self.TreatmentOptions.get()}") #OK - just for log
            # current_treatment_char = current_treatment.split(" ")[1]
            logger.debug(f"current_treatment_char = {current_condition_char}")
            THE_HISTORY.fish_adder(project_name=self.CURRENT_PROJECT, 
                                 test_num=current_test_index,
                                 batch_num=current_batch_index, 
                                 treatment_num=current_condition_char, 
                                 target_amount=target_amount, 
                                 task=task,
                                 modify_history=False)
        elif treatment_mode == "all":
            logger.debug("Modify folders for all treatment")
            THE_HISTORY.fish_adder(project_name=self.CURRENT_PROJECT, 
                                 test_num=current_test_index,
                                 batch_num=current_batch_index, 
                                 treatment_num=treatment_mode,
                                 target_amount=target_amount, 
                                 task=task,
                                 modify_history=False)


    def nk_add(self):
        logger.debug("nk_add_button pressed")
        target_amount = self.nested_key_1_frame.add_entry()
        if self.nested_key_2_frame.null_label == None:
            self.nested_key_2_frame.add_entry()
        self.folder_changer(target_amount, task="add")

    def nk_remove(self):
        logger.debug("nk_remove_button pressed")
        target_amount = self.nested_key_1_frame.remove_entry()
        if self.nested_key_2_frame.null_label == None:
            self.nested_key_2_frame.remove_entry()
        self.folder_changer(target_amount, task="remove")

    # def nk1_add(self):
    #     logger.debug("nk1_add_button pressed")
    #     target_amount = self.nested_key_1_frame.add_entry()
    #     self.folder_changer(target_amount)

    # def nk2_add(self):
    #     logger.debug("nk2_add_button pressed")
    #     target_amount = self.nested_key_2_frame.add_entry()
    #     self.folder_changer(target_amount)

    # def nk1_remove(self):
    #     logger.debug("nk1_remove_button pressed")
    #     target_amount = self.nested_key_1_frame.remove_entry()
    #     self.folder_changer(target_amount)

    # def nk2_remove(self):
    #     logger.debug("nk2_remove_button pressed")
    #     target_amount = self.nested_key_2_frame.remove_entry()
    #     self.folder_changer(target_amount)
        
        
    def import_video(self):

        try:
            project_dir = Path(THE_HISTORY.get_project_dir(self.CURRENT_PROJECT))
        except TypeError:
            message_ = "Please select a project before using Import Videos function"
            tkinter.messagebox("Warning", message_)

        video_add_window = VideoAdd(self, project_dir, list2=self.CONDITIONLIST, list3=self.BATCHLIST)
        

    def access_history(self, command_type, batch_name=None, edit_command=None):
        logger.debug("Accessing history file")

        # load the history file
        try:
            with open(HISTORY_PATH, "r") as file:
                projects_data = json.load(file)
        except:
            ErrorType = "Empty history file"
            logger.warning(ErrorType)
            return None, ErrorType

        # current project name
        cp = self.CURRENT_PROJECT

        # Check if the project exists
        if cp not in projects_data.keys():
            ErrorType = "Project doesn't exist"
            logger.warning(ErrorType)
            return None, ErrorType
    
        # How many batch files are there?
        batch_quantity = 0
        batch_list = []
        for key in projects_data[cp].keys():
            if "Batch" in key:
                batch_quantity += 1
                batch_list.append(key)

        if batch_quantity == 0:
            ErrorType = "No batches"
            logger.warning(ErrorType)
            return None, ErrorType

        # Modify the history file
        if command_type == "add":
            logger.debug("Command = add")
            if batch_name in projects_data[cp].keys():
                ErrorType = "Batch already exists, can't add"
                logger.warning(ErrorType)
                return None, ErrorType
            else:
                example_key = list(projects_data[cp].keys())[0]
                projects_data[cp][batch_name] = projects_data[cp][example_key]
                with open(HISTORY_PATH, "w") as file:
                    json.dump(projects_data, file, indent=4)
                return None, None
            
        elif command_type == "remove":
            logger.debug("Command = remove")
            if batch_quantity == 1:
                ErrorType = "Last batch, can't remove"
                logger.warning(ErrorType)
                return None, ErrorType
            elif batch_name not in projects_data[cp].keys():
                ErrorType = f"{batch_name} doesn't exist"
                logger.warning(ErrorType)
                return None, ErrorType
            else:
                # Remove the batch
                projects_data[cp].pop(batch_name)
                with open(HISTORY_PATH, "w") as file:
                    json.dump(projects_data, file, indent=4)
                return None, None
            
        elif command_type == "edit":
            logger.debug("Command = edit")
            if edit_command == None:
                ErrorType = "No edit command given"
                logger.warning(ErrorType)
                return None, ErrorType
            else:
                treatment = edit_command[0]
                value_pos = edit_command[1]
                new_value = edit_command[2]
                try:
                    projects_data[cp][batch_name][treatment][value_pos] = new_value
                    with open(HISTORY_PATH, "w") as file:
                        json.dump(projects_data, file, indent=4)
                    return None, None
                except:
                    logger.error("Invalid edit command")
                    raise Exception("Invalid edit command")

        elif command_type == "load batch list":
            logger.debug("Command = load batch list")
            return batch_list, None
        
        elif command_type == "load treatment list":
            logger.debug("Command = load treatment list")
            logger.debug(f"CP: {cp} ,Batch name: {batch_name}")
            treatments = []
            for treatment_key in projects_data[cp][batch_name].keys():
                _name = projects_data[cp][batch_name][treatment_key][0]
                _dose = projects_data[cp][batch_name][treatment_key][1]
                _unit = projects_data[cp][batch_name][treatment_key][2]
                if _unit == "":
                    treatments.append(_name)
                else:
                    treatments.append(f"{_name} {_dose} {_unit}")
            logger.debug(f"Treatments: {treatments}")
            return treatments, None
        
        else:
            logger.error("Invalid command type")
            raise Exception("Invalid command type")

    def add_batch(self):
        logger.debug("Adding batch")
        new_batch_num = len(self.BATCHLIST) + 1
        self.BATCHLIST.append("Batch " + str(new_batch_num))

        # Update the batch options
        self.BatchOptions.configure(values=self.BATCHLIST)

        # Set the batch to the last batch
        self.BatchOptions.set(self.BATCHLIST[-1])

        # Modify history file
        _, ErrorType = self.access_history("add", f"Batch {new_batch_num}")

        if ErrorType != None:
            logger.error(ErrorType)
            tkinter.messagebox.showerror("Error", ErrorType)
            return

        # Create new batch directories and hyp files
        self.save_project(batch_num = new_batch_num, subsequent_save = True)


    def remove_batch(self):
        logger.debug("Removing batch")
        selected_batch = self.BatchOptions.get()

        # Pop-up window to confirm deletion
        if not tkinter.messagebox.askokcancel("Delete Batch", f"Are you sure you want to delete {selected_batch}?"):
            return

        # Modify history file
        _, ErrorType = self.access_history("remove", selected_batch)

        if ErrorType != None:
            logger.error(ErrorType)
            tkinter.messagebox.showerror("Error", ErrorType)
            return

        self.BATCHLIST, _ = self.access_history("load batch list")

        # Update the batch options
        self.BatchOptions.configure(values=self.BATCHLIST)

        # Set the batch to the last batch
        self.BatchOptions.set(self.BATCHLIST[-1])

        # Remove the batch directories and hyp files
        self.delete_batch(selected_batch)

    
    def delete_batch(self, batch_name):
        logger.debug("Deleting batch")

        batch_num = batch_name.split(" ")[1]

        batch_ord = ORDINALS[int(batch_num)-1]

        project_dir = Path(THE_HISTORY.get_project_dir(self.CURRENT_PROJECT))

        # Find all directory in project_dir, at any level, that contain batch_ord, use shutil.rmtree to delete them
        for dir in project_dir.glob(f"**/*{batch_ord}*"):
            shutil.rmtree(dir)

        # Delete the hyp file
        batch_static_dir = project_dir / "static" / f"Batch {batch_num}"
        shutil.rmtree(batch_static_dir)


    def param_display(self, selected_test = None, condition = "A", batch_num = "1"):

        self.parameters_frame.load_parameters(project_name = self.CURRENT_PROJECT, selected_task = selected_test, condition=condition, batch_num=batch_num, nested_key = 0)
        nested_key_1 = self.nested_key_1_frame.load_parameters(project_name = self.CURRENT_PROJECT, selected_task = selected_test, condition=condition, batch_num=batch_num, nested_key = 1)
        nested_key_2 = self.nested_key_2_frame.load_parameters(project_name = self.CURRENT_PROJECT, selected_task = selected_test, condition=condition, batch_num=batch_num, nested_key = 2)

        if self.nested_key_1_frame.null_label != None:
            # hide add & remove button
            self.nk_add_button.hide()
            self.nk_remove_button.hide()
            # self.nk_import_button.hide()
            self.nk_selector_button.hide()
        else:
            self.nk_add_button.show()
            self.nk_remove_button.show()
            # self.nk_import_button.show()
            self.nk_selector_button.show()

        # if self.nested_key_2_frame.null_label != None:
        #     # hide add & remove button
        #     self.nk2_add_button.hide()
        #     self.nk2_remove_button.hide()
        # else:
        #     self.nk2_add_button.show()
        #     self.nk2_remove_button.show()

        self.LoadedProject.configure(text=self.CURRENT_PROJECT)
        self.nested_key_1_header.configure(text=nested_key_1)
        self.nested_key_2_header.configure(text=nested_key_2)

        self.PREVIOUS_TEST = selected_test
        self.PREVIOUS_BATCH = batch_num
        self.PREVIOUS_CONDITION = condition


    def update_param_display(self, event=None, load_type="not_first_load"):
        assert load_type in ["not_first_load", "first_load"]

        if load_type == "first_load":
            logger.info("Initial display of parameters, batch = 1, test = first test, condition = A")

            selected_test = self.TESTLIST[0]

            #set TestOptions to the first test
            self.TestOptions.set(selected_test)
            #set TreatmentOptions to the first condition
            self.TreatmentOptions.set(self.CONDITIONLIST[0])
            #set BatchOptions to the first batch
            self.BatchOptions.set(self.BATCHLIST[0])

            self.param_display(selected_test = selected_test)
            return
        
        self.save_parameters(mode = "previous")
        logger.debug("Saved the previous parameters")

        batch_num = self.BatchOptions.get().split()[1]
        logger.debug(f"Batch DropDown: {self.PREVIOUS_BATCH} -> {batch_num}")
        selected_test = self.TestOptions.get()
        logger.debug(f"Test DropDown: {self.PREVIOUS_TEST} -> {selected_test}")
        
        condition = self.TreatmentOptions.get() #OK - just for log
        logger.debug(f"Condition DropDown: {self.PREVIOUS_CONDITION} -> {condition}")
        # convert condition_index to letter 1 -> A
        current_condition_char = self.get_treatment_char()

        self.param_display(selected_test = selected_test, condition = current_condition_char, batch_num = batch_num)


    def save_parameters(self, mode = "current", treatment_mode="current"):
        assert mode in ["current", "previous"]

        if mode == "current":
            logger.debug("Save button pressed, save the current parameters")
            # Get the selected test type
            selected_test = self.TestOptions.get()
            batch_num = self.BatchOptions.get().split()[1]
            condition = self.get_treatment_char()
            logger.debug(f"Condition: {condition}")
        else:
            logger.debug("Other option selected, save the previous parameters")
            selected_test = self.PREVIOUS_TEST
            batch_num = self.PREVIOUS_BATCH
            condition = self.PREVIOUS_CONDITION
            logger.debug(f"Condition: {condition}")

        # Save the parameters
        # save_parameters(self, project_name, selected_task, condition, batch_num, mode = 'single'):
        
        self.parameters_frame.save_parameters(project_name = self.CURRENT_PROJECT, selected_task = selected_test, condition=condition, batch_num=batch_num, treatment_mode = treatment_mode)
        self.nested_key_1_frame.save_parameters(project_name = self.CURRENT_PROJECT, selected_task = selected_test, condition=condition, batch_num=batch_num, treatment_mode = treatment_mode)
        self.nested_key_2_frame.save_parameters(project_name = self.CURRENT_PROJECT, selected_task = selected_test, condition=condition, batch_num=batch_num, treatment_mode = treatment_mode)


    ### DELETE PROJECT BUTTON FUNCTION ###
    def delete_project(self):

        # create confirmation box
        choice = tkinter.messagebox.askquestion("Delete Project", "Are you sure you want to delete this project?")
        if choice == "no":
            return

        # Get the selected project
        selected_project = self.scrollable_frame.get_selected_project()

        if selected_project == "":
            tkinter.messagebox.showerror("Error", "Please select a project")
            return

        # Delete the project from the history file
        with open(HISTORY_PATH, "r") as file:
            projects_data = json.load(file)

        project_dir = projects_data[selected_project]["DIRECTORY"]

        # Delete the project directory
        try:
            shutil.rmtree(project_dir)
            logger.info(f"Deleted project directory: {project_dir}")
        except:
            logger.debug("Project directory does not exist: , just remove from History")

        del projects_data[selected_project]

        with open(HISTORY_PATH, "w") as file:
            json.dump(projects_data, file, indent=4)

        self.CURRENT_PROJECT = ""

        logger.info("Set current project to blank")

        # Refresh the project list
        logger.debug("Refresh projects")
        self.refresh_projects()

        # Refresh the project details
        self.refresh_projects_detail()

    ### LOAD PROJECT BUTTON FUNCTION ###

    def refresh_projects_detail(self):
        logger.debug("Refresh projects detail")

        # Clear existing project details labels
        self.project_detail_container.clear()

        # Reload the project details
        self.project_detail_container.load_project_details(self.CURRENT_PROJECT)
        

    def load_project(self, custom_project=None):
        logger.debug("Load project button pressed")

        if custom_project == None:
            selected_project = self.scrollable_frame.get_selected_project()
        else:
            selected_project = custom_project
            # set the current project to the custom project
            self.scrollable_frame.set_selected_project(custom_project)

        self.CURRENT_PROJECT = selected_project

        logger.info(f"Current project: {self.CURRENT_PROJECT}")

        # Update the batch options
        self.BATCHLIST, ErrorType = self.access_history("load batch list")
        if ErrorType != None:
            tkinter.messagebox.showerror("Error", ErrorType)
            return
        
        
        self.BatchOptions.configure(values=self.BATCHLIST)

        retry = 0 
        while retry<3:
            try:
                self.CONDITIONLIST, ErrorType = self.access_history("load treatment list", batch_name = self.BatchOptions.get())
                logger.debug("Loaded condition list")
                logger.debug(f"Possible warning: {ErrorType}")
                break
            except:
                logger.warning(f"Batch {self.BatchOptions.get()} does not exist in this project, try another batch")
                self.BatchOptions.set(self.BATCHLIST[0])
                retry += 1
                logger.debug(f"Retried {retry} times")
        else:
            logger.error("Failed to load condition list, please check the project directory")
            tkinter.messagebox.showerror("Error", "Failed to load condition list, please check the project directory")
            return

        if ErrorType != None:
            tkinter.messagebox.showerror("Error", ErrorType)
            return
        
        #set values of TreatmentOptions
        self.TreatmentOptions.configure(values=self.CONDITIONLIST)
        #set current value to first choice
        self.TreatmentOptions.set(self.CONDITIONLIST[0])

        self.refresh_projects_detail()

        self.update_param_display(load_type = "first_load")



    ### CREATE PROJECT BUTTON FUNCTION ###

    def create_project(self):
        logger.debug("Create project button pressed")

        self.PROJECT_CREATED = False
        self.project_input_window()

        if self.PROJECT_CREATED:

            with open(HISTORY_PATH, "r") as file:
                projects_data = json.load(file)

            # latest_project = list(projects_data.keys())[-1]
            print(f"Project name: {self.CURRENT_PROJECT}")
            for treatment, details in projects_data[self.CURRENT_PROJECT].items():
                print(f"{treatment}: {details}")

            self.refresh_projects()
            # select the newly created project in the list
            self.scrollable_frame.select_project(self.CURRENT_PROJECT)

            self.save_project()

            self.load_project(custom_project=self.CURRENT_PROJECT)

        else:
            print("Project not created")


    def refresh_projects(self):
        logger.debug("Refresh projects")

        # Clear existing project labels
        self.scrollable_frame.clear_projects()

        # Read the projects.json file and add project names to the list
        try:
            with open(HISTORY_PATH, "r") as file:
                projects_data = json.load(file)
        except:
            print("No projects found or no record of projects")
            return

        for project_name in projects_data.keys():
            self.scrollable_frame.add_project(project_name)


    def directories_maker(self, project_dir, batch_num, subsequent_save = False):
        logger.debug(f"Make directories for {project_dir}, batch_num = {batch_num}")

        if os.path.exists(project_dir):
            if not subsequent_save:
                return
            else:
                pass
        else:
            os.mkdir(project_dir)

        # turn batch_num into ordinal
        batch = ORDINALS[int(batch_num)-1]

        # change number 2 into B, using mathematically

        with open(HISTORY_PATH, "r") as file:
            projects_data = json.load(file)

        if not subsequent_save:
            treatments = projects_data[self.CURRENT_PROJECT]["Batch 1"]
        else:
            treatments = projects_data[self.CURRENT_PROJECT][f"Batch {batch_num}"]

        all_paths = {}

        all_paths['Parent'] = []
        for i, test in enumerate(self.TESTLIST):
            temp = f"0{i+1} - {test}"
            all_paths['Parent'].append(temp)
        
        normal_parents = []
        group_parents = []
        for path in all_paths['Parent']:
            if "Shoaling" in path:
                group_parents.append(path)
            else:
                normal_parents.append(path)

        Shoaling_Tank_Count_Dict = {}
        Other_Tank_Count_Dict = {}

        for k, v in treatments.items():
            shoaling_tank_count = 0
            other_tank_count = 0
            char = k.split()[1]
            if char == "A":
                tail = f"A - Control ({batch} Batch)"
            else:
                # check type of v[1]
                if float(v[1]) >= 10:
                    v[1] = str(int(float(v[1])))
                else:
                    v[1] = str(float(v[1]))
                tail = f"{char} - {v[0]} {v[1]} {v[2]} ({batch} Batch)"

            fish_num = int(v[3])
            if fish_num > other_tank_count:
                other_tank_count = fish_num
            all_paths[f"Child-{char}"] = [f"{parent}\\{tail}" for parent in all_paths['Parent']]
            for i in range(1, fish_num+1):
                all_paths[f"Child-{char}"].extend([f"{parent}\\{tail}\\{i}" for parent in normal_parents])
            
            fish_group = fish_num // 3
            if fish_group > shoaling_tank_count:
                shoaling_tank_count = fish_group
            for i in range(1, fish_group+1):
                all_paths[f"Child-{char}"].extend([f"{parent}\\{tail}\\{i}" for parent in group_parents])

            Shoaling_Tank_Count_Dict[char] = shoaling_tank_count
            Other_Tank_Count_Dict[char] = other_tank_count

        for k, v in all_paths.items():
            if k == "Parent" and subsequent_save:
                continue
            for path in v:
                os.makedirs(os.path.join(project_dir, path))
            
        return Shoaling_Tank_Count_Dict, Other_Tank_Count_Dict, len(treatments)


    def save_project(self, batch_num = 1, subsequent_save = False):
        logger.info(f"Save project {self.CURRENT_PROJECT}")

        if not subsequent_save:
            save_dir = tkinter.filedialog.askdirectory()
            save_dir = Path(save_dir)
            project_dir = save_dir / self.CURRENT_PROJECT
        else:
            project_dir = Path(THE_HISTORY.get_project_dir(self.CURRENT_PROJECT))

        # project_dir.mkdir(parents=True, exist_ok=True)

        shoaling_tank_count_dict, other_tank_count_dict, treatment_count = self.directories_maker(project_dir, batch_num, subsequent_save)

        with open(HISTORY_PATH, "r") as file:
            projects_data = json.load(file)
        
        # save the directory of the project to the projects_data
        projects_data[self.CURRENT_PROJECT]["DIRECTORY"] = str(project_dir)

        with open(HISTORY_PATH, "w") as file:
            json.dump(projects_data, file, indent=4)

        #clone the hyp_{}.json file to static_path
        # ori_static_path = Path(__file__).parent / "Bin"
        temp_path = Path(__file__).parent / "temp"

        # make temp_path if not exist
        if not temp_path.exists():
            temp_path.mkdir()

        # Then copy to the project_static_dirs
        project_static_dirs = get_static_dir(project_dir, batch_num=batch_num, treatment_count = treatment_count)

        for treatment_char in shoaling_tank_count_dict.keys():
            shoaling_tank_count = shoaling_tank_count_dict[treatment_char]
            other_tank_count = other_tank_count_dict[treatment_char]
            project_static_dir = project_static_dirs[treatment_char]

            logger.debug("For treatment %s, shoaling_tank_count = %d, other_tank_count = %d", treatment_char, shoaling_tank_count, other_tank_count)
            logger.debug(f"Static dir for treatment {treatment_char} is {project_static_dir}")

            # take all the files start with "hyp_" and end with ".json" from ori_static_path
            # and copy them to a temp_path for modification first
            for file_name, value in CONSTANTS.items():
                file_path = temp_path / file_name
                if file_path.exists():
                    # remove the file if it already exists
                    file_path.unlink()
                with open(file_path, "w") as f:
                    json.dump(value, f, indent=4)

            for file in temp_path.glob("hyp_*.json"):
                if "shoaling" in file.name:
                    desired_key_num = shoaling_tank_count
                else:
                    desired_key_num = other_tank_count

                with open(file, "r") as f:
                    data = json.load(f)
                for value in data.values():
                    if not isinstance(value, dict):
                        continue

                    if len(value) < desired_key_num:
                        logger.debug("Tank number is not enough, adding extra tanks")
                        # check type of value.values(), if it is a list, new_value = [0,0], else new_value = 0
                        if type(list(value.values())[0]) == list:
                            default_value = [0,0]
                        else:
                            default_value = 0
                        while len(value) < desired_key_num:
                            temp_key = len(value)+1
                            logger.debug(f"Adding tank: {temp_key} with default value: {default_value} to file {file.name}")
                            value[temp_key] = default_value
                    elif len(value) > desired_key_num:
                        logger.debug("Tank number is too much, removing extra tanks")
                        while len(value) > desired_key_num:
                            logger.debug("Removing tank")
                            value.popitem()
                    else:
                        pass

                with open(file, "w") as f:
                    json.dump(data, f, indent=4)

            project_static_dir.mkdir(parents=True, exist_ok=True)

            for file in temp_path.glob("hyp_*.json"):
                shutil.copy(file, project_static_dir)
                logger.debug(f"Copying {file.name} to {project_static_dir}")

            # Remove hyp files from temp
            for file in temp_path.glob("hyp_*.json"):
                os.remove(file)


    def project_input_window(self):
        logger.info("Project input window opened")

        batch_name = "Batch 1"

        treatment_widgets = []

        bold_font = customtkinter.CTkFont(size = 15, weight="bold")

        def add_treatment():
            logger.debug("Add treatment button clicked")

            treatment_row = len(treatment_widgets)*3 + r + 1
            treatment_name = f"Treatment {chr(ord('C') + len(treatment_widgets))}:"

            treatment_label = customtkinter.CTkLabel(top_canvas, text=treatment_name, font=bold_font)
            treatment_label.grid(row=treatment_row, column=0, pady=(20, 5))
            treatment_entry = customtkinter.CTkEntry(top_canvas)
            treatment_entry.grid(row=treatment_row, column=1, pady=(20, 5))

            dose_label = customtkinter.CTkLabel(top_canvas, text="Dose:")
            dose_label.grid(row=treatment_row + 1, column=0, pady=5)
            dose_entry = customtkinter.CTkEntry(top_canvas)
            dose_entry.grid(row=treatment_row + 1, column=1, pady=5)
            unit_optionmenu = customtkinter.CTkOptionMenu(top_canvas, values=["ppm", "ppb"])
            unit_optionmenu.grid(row=treatment_row + 1, column=2, pady=5)

            fish_number_label = customtkinter.CTkLabel(top_canvas, text="Fish Number:")
            fish_number_label.grid(row=treatment_row + 2, column=0, pady=5)
            fish_number_entry = customtkinter.CTkEntry(top_canvas)
            fish_number_entry.grid(row=treatment_row + 2, column=1, pady=5)

            treatment_widgets.append((treatment_entry, dose_entry, unit_optionmenu, fish_number_entry))

        def get_values():
            project_name = project_name_entry.get()
            self.CURRENT_PROJECT = project_name
            try:
                note = treatment_a_entry.get()
            except:
                note = ""
            try:
                treatment_list = {
                    "Treatment A": [
                        "Control",
                        0,
                        "",
                        int(fish_number_a_entry.get()),
                        note
                    ],
                    "Treatment B": [
                        treatment_b_entry.get(),
                        float(dose_b_entry.get()),
                        unit_b_optionmenu.get(),
                        int(fish_number_b_entry.get()),
                        note
                    ]
                }
            except Exception as e:
                #show message box of error
                print(e)
                tkinter.messagebox.showerror("Error", "Please fill the required fields with right type of value")

            for i, (treatment_entry, dose_entry, unit_optionmenu, fish_number_entry) in enumerate(treatment_widgets):
                treatment_name = f"Treatment {chr(ord('C') + i)}"
                treatment_list[treatment_name] = [
                    treatment_entry.get(),
                    float(dose_entry.get()),
                    unit_optionmenu.get(),
                    int(fish_number_entry.get()),
                    note
                ]

            # Save values to projects.json
            project_data = {
                project_name: {
                    batch_name : treatment_list
                    }
                }

            try:
                with open(HISTORY_PATH, "r") as file:
                    existing_data = json.load(file)
                if project_name in existing_data:
                    # Display message box of error
                    tkinter.messagebox.showerror("Error", "Project already exists")
                else:
                    existing_data.update(project_data)
                    self.PROJECT_CREATED = True
                    with open(HISTORY_PATH, "w") as file:
                        json.dump(existing_data, file, indent=2)
                    input_window.destroy()  # Move this line inside the else block
            except:
                existing_data = project_data
                self.PROJECT_CREATED = True
                with open(HISTORY_PATH, "w") as file:
                    json.dump(existing_data, file, indent=2)
                input_window.destroy()  # Move this line inside the except block

        def cancel_button_command():
            logger.debug("Cancel button clicked")

            self.PROJECT_CREATED = False
            input_window.destroy()


        input_window = tkinter.Toplevel(self)
        # set window size
        input_window.geometry("400x500")

        input_window.title("Project Input")

        # Top Canvas
        top_canvas = customtkinter.CTkScrollableFrame(input_window, width = 380)
        # expand the canvas to fill the window
        input_window.rowconfigure(0, weight=1)
        top_canvas.grid(row=0, column=0, sticky="nsew")

        r=0
        # Project name
        project_name_label = customtkinter.CTkLabel(top_canvas, text="Project name:", font=bold_font)
        project_name_label.grid(row=r, column=0, pady=5)
        project_name_entry = customtkinter.CTkEntry(top_canvas)
        project_name_entry.grid(row=r, column=1, pady=5)

        r+=1
        # Treatment A (Control)
        treatment_a_label = customtkinter.CTkLabel(top_canvas, text="Treatment A:", font=bold_font)
        treatment_a_label.grid(row=r, column=0, pady=5)
        treatment_a_entry = customtkinter.CTkEntry(top_canvas)
        treatment_a_entry.grid(row=r, column=1, pady=5)

        hover_button = tkinter.Button(top_canvas, text="?")
        hover_button.grid(row=r, column=2, pady=5)
        CreateToolTip(hover_button, text = 'Control condition\n'
                 'Leave blank if you used pure water\n'
                 'The info you put here would be saved as Note\n'
        )

        r+=1
        # Fish number
        fish_number_a_label = customtkinter.CTkLabel(top_canvas, text="Fish Number:")
        fish_number_a_label.grid(row=r, column=0, pady=5)
        fish_number_a_entry = customtkinter.CTkEntry(top_canvas)
        fish_number_a_entry.grid(row=r, column=1, pady=5)
        
        r+=1
        # Treatment B
        treatment_b_label = customtkinter.CTkLabel(top_canvas, text="Treatment B:", font=bold_font)
        treatment_b_label.grid(row=r, column=0, pady=(20, 5))
        treatment_b_entry = customtkinter.CTkEntry(top_canvas)
        treatment_b_entry.grid(row=r, column=1, pady=(20, 5))

        r+=1
        # Dose
        dose_label = customtkinter.CTkLabel(top_canvas, text="Dose:")
        dose_label.grid(row=r, column=0, pady=5)
        dose_b_entry = customtkinter.CTkEntry(top_canvas)
        dose_b_entry.grid(row=r, column=1, pady=5)
        unit_b_optionmenu = customtkinter.CTkOptionMenu(top_canvas, values=["ppm", "ppb"])
        unit_b_optionmenu.grid(row=r, column=2, pady=5)

        r+=1
        # Fish number
        fish_number_b_label = customtkinter.CTkLabel(top_canvas, text="Fish Number:")
        fish_number_b_label.grid(row=r, column=0, pady=5)
        fish_number_b_entry = customtkinter.CTkEntry(top_canvas)
        fish_number_b_entry.grid(row=r, column=1, pady=5)

        # Bottom Canvas
        bottom_canvas = customtkinter.CTkFrame(input_window)
        bottom_canvas.grid(row=1, column=0, sticky="nsew")

        # Add button
        add_button = customtkinter.CTkButton(bottom_canvas, text="Add Treatment", 
                                             command=add_treatment)
        add_button.grid(row=0, column=0, padx=5, pady=20)

        # Confirm button
        confirm_button = customtkinter.CTkButton(bottom_canvas, text="CONFIRM", 
                                                 font = bold_font,
                                                 command=get_values)
        confirm_button.grid(row=1, column=0, padx=5, pady=20)

        # Cancel button
        cancel_button = customtkinter.CTkButton(bottom_canvas, text="CANCEL", 
                                                font = bold_font,
                                                command=cancel_button_command)
        cancel_button.grid(row=1, column=1, padx=5, pady=20)

        input_window.wait_window()


    def set_state(self, event=None, set_project=None, set_batch=None, set_test=None, set_treatment=None):
        if set_project != None:
            try:
                self.load_project(custom_project=set_project)
            except:
                logger.warning("Failed to load Project {}".format(set_project))

        if set_batch != None:
            try:
                self.BatchOptions.set(set_batch)
                self.update_param_display()
            except:
                logger.warning("Failed to load Batch {}".format(set_batch))

        if set_test != None:
            try:
                self.TestOptions.set(set_test)
                self.update_param_display()
            except:
                logger.warning("Failed to load Test {}".format(set_test))

        if set_treatment != None:
            try:
                self.TreatmentOptions.set(set_treatment)
                self.update_param_display()
            except:
                logger.warning("Failed to load Treatment {}".format(set_treatment))        

    def mismatch_show(self, treatment_name):
        self.set_state(set_treatment=treatment_name)

        test_name = self.TestOptions.get()
        batch_num = self.BatchOptions.get() # Batch 1
        batch_num = int(batch_num.split(" ")[1]) # 1
        treatment_char = get_treatment_char_from_name(treatment_name, self.CONDITIONLIST)

        treatment_path = Path(THE_HISTORY.get_treatment_dir(self.CURRENT_PROJECT, 
                                                            test_name, 
                                                            batch_num, 
                                                            treatment_char))
        # use window explorer to open the folder
        logger.debug("Open folder: {}".format(treatment_path))
        try:
            os.startfile(treatment_path)
        except:
            logger.warning("Failed to open folder: {}".format(treatment_path))

    ### ANALYZE BUTTON ###

    def create_progress_window(self, title = "Analysis Progress"):
        progress_window = tkinter.Toplevel(self)
        progress_window.title(title)
        progress_window.geometry("300x100")

        progress_label = tkinter.Label(progress_window, text="Analyzing...", font=('Helvetica', 12))
        progress_label.pack(pady=(10, 0))

        progress_bar = ttk.Progressbar(progress_window, mode='determinate', length=200)
        progress_bar.pack(pady=(10, 20))

        # progress_window.protocol("WM_DELETE_WINDOW", self.set_stop_thread) # Set protocol

        return progress_bar
    

    # def set_stop_thread(self):
    #     self.stop_thread = True
    

    def analyze_project(self):
        logger.info("Start analyzing project")
        self.stop_thread = False

        if self.CURRENT_PROJECT == "":
            tkinter.messagebox.showerror("Error", "Please select a project")
            return
        
        # save the current parameters
        self.save_parameters(mode='current')

        project_dir = Path(THE_HISTORY.get_project_dir(self.CURRENT_PROJECT))

        # get selected task
        test_name = self.TestOptions.get()

        try:
            BATCH_NUMBER = int(self.BatchOptions.get().split()[1])
        except ValueError:
            BATCH_NUMBER = 1

        overwrite = False

        while True:

            progress_bar = self.create_progress_window()
            total_time, notification, ERROR = autoanalyzer(PROJECT_DIR = project_dir, 
                                                           BATCHNUM = BATCH_NUMBER, 
                                                           TASK = test_name, 
                                                           PROGRESS_BAR = progress_bar, 
                                                           OVERWRITE = overwrite, 
                                                           skip_list = self.SKIP_DICT)

            if ERROR == None:
                progress_bar.master.destroy()  # Close the progress window
                display_info = f"{notification}.\n Total time taken: {total_time} seconds"
                tkinter.messagebox.showinfo("Analysis Complete", display_info)
                logger.info(display_info)
                break
            elif ERROR == "Existed":
                logger.debug("File existed")
                # pop up a window to ask if the user wants to overwrite
                overwrite = tkinter.messagebox.askyesno("Warning", f"{notification}. Do you want to overwrite?")
                if overwrite:
                    logger.debug("User chose to overwrite")
                    progress_bar.master.destroy()
                    continue
                else:
                    logger.debug("User chose not to overwrite")
                    progress_bar.master.destroy()
                    break
            elif ERROR == "File Opened":
                logger.debug("File opened")
                message = f"{notification}. \n Please close the file and press OK to proceed. \n Press No to cancel."
                proceed = tkinter.messagebox.askyesno("Error", message)
                if proceed:
                    progress_bar.master.destroy()
                    continue
                else:
                    progress_bar.master.destroy()
                    break
            elif ERROR == "Mismatched":
                logger.debug("Mismatched")
                treatment_char, required_params, current_params = notification.split(";")
                treatment_name = get_treatment_name_from_char(treatment_char, self.CONDITIONLIST)
                message = f"In {treatment_name}, current parameters = {current_params} mismatched with fish numbers of {required_params}. \n Press 'Go' to go to change the parameters"
                progress_bar.master.destroy()
                _ = CustomDialog(self, title="Mismatched", message=message, button_text="Go", button_command= lambda : self.mismatch_show(treatment_name=treatment_name))
                break

    def trajectories_filler(self, filler_path, ref_df, ref_tanks, input=True):
        # create a box to input X and Y coordinates
        input_window = tkinter.Toplevel(self)
        input_window.title("Input Filler Coordinates")
        input_window.geometry("600x120")
        #move the window to the center of the screen, bring it to front
        input_window.lift()
        input_window.geometry(f"600x160+{int(input_window.winfo_screenwidth()/2 - 600/2)}+{int(input_window.winfo_screenheight()/2 - 160/2)}")

        #[TODO] MAKE IT INTO A CHOOSE ON SCREEN THING

        # Top Canvas
        top_canvas = customtkinter.CTkFrame(input_window)
        top_canvas.grid(row=0, column=0, sticky="nsew")

        ref_df_length = len(ref_df)

        # df_len_label = customtkinter.CTkLabel(top_canvas, text=f"Total number of frames")
        # df_len_label.grid(row=0, column=0, columnspan=2, padx=5, pady=10)
        # df_len_entry = customtkinter.CTkEntry(top_canvas)
        # df_len_entry.grid(row=0, column=2, columnspan=2, padx=5, pady=10)
        # df_len_entry.insert(0, ref_df_length)

        x_labels = {}
        y_labels = {}
        x_entries = {}
        y_entries = {}
        for i in ref_tanks:
            x_labels[i] = customtkinter.CTkLabel(top_canvas, text=f"Tank {i} X:")
            x_labels[i].grid(row=i, column=0, padx=5, pady=5)
            x_entries[i] = customtkinter.CTkEntry(top_canvas)
            x_entries[i].grid(row=i, column=1, padx=5, pady=5)



            y_labels[i] = customtkinter.CTkLabel(top_canvas, text=f"Tank {i} Y:")
            y_labels[i].grid(row=i, column=2, padx=5, pady=5)
            y_entries[i] = customtkinter.CTkEntry(top_canvas)
            y_entries[i].grid(row=i, column=3, padx=5, pady=5)

            selector_button = customtkinter.CTkButton(top_canvas, text="Select", command=lambda i=i: Selector(master=self, 
                                                                                                              x_entry=x_entries[i],
                                                                                                              y_entry=y_entries[i]))
            selector_button.grid(row=i, column=4, padx=5, pady=5)


        def make_filler_coordinates():
            output_df = pd.DataFrame()
            try:
                # current_df_length = int(float(df_len_entry.get()))
                current_df_length = ref_df_length
            except ValueError:
                tkinter.messagebox.showerror("Error", "Please enter a valid number in the Total number of frames box")
                return
            
            for i in ref_tanks:
                try:
                    x = float(x_entries[i].get())
                    y = float(y_entries[i].get())
                except ValueError:
                    tkinter.messagebox.showerror("Error", "Please enter valid numbers")
                    return
                output_df[f"X{i}"] = [x] * current_df_length
                output_df[f"Y{i}"] = [y] * current_df_length

            separator = "\t"
            output_df.to_csv(filler_path, index=False, sep=separator)

            #create a message box saying that the coordinates have been saved
            tkinter.messagebox.showinfo("Success", f"Filler coordinates have been saved to: \n{filler_path}")

            # close the window
            input_window.destroy()
            return True
        
        def canceling():
            input_window.destroy()
            return False

        # Bottom Canvas
        bottom_canvas = customtkinter.CTkFrame(input_window)
        bottom_canvas.grid(row=1, column=0, sticky="nsew")


        confirm_button = customtkinter.CTkButton(bottom_canvas, text="CONFIRM",
                                                    command=make_filler_coordinates)
        confirm_button.grid(row=0, column=0, padx=15, pady=20)

        cancel_button = customtkinter.CTkButton(bottom_canvas, text="CANCEL",
                                                command=canceling)
        cancel_button.grid(row=0, column=1, padx=15, pady=20)

        self.wait_window(input_window)



    def pre_analyze_check(self):
        logger.debug("Start pre-analyze check")

        if self.CURRENT_PROJECT == "":
            tkinter.messagebox.showerror("Error", "Please select a project")
            return False

        if self.TestOptions.get() == "":
            tkinter.messagebox.showerror("Error", "Please select a task")
            return False

        if self.BatchOptions.get() == "":
            tkinter.messagebox.showerror("Error", "Please select a batch")
            return False
        
        # CHECK TRAJECTORIES VALIDITY
        status, false_keys_dict, treatments_checker_dict = self.trajectories_check(mode='all')

        if status == "All True":
            logger.debug("All trajectories are valid")
            pass
        elif status == "All False":
            logger.warning("Tried to analyze project with no valid trajectories")
            tkinter.messagebox.showerror("Error", "Current task has no valid trajectories")
            return False
        else:
            logger.warning("Some trajectories are invalid")
            skip_dict = {}
            for treatment_char in false_keys_dict.keys():
                false_keys = false_keys_dict[treatment_char]
                checker_dict = treatments_checker_dict[treatment_char]

                logger.warning(f"Trajectories in treatment {treatment_char} is invalid.")
                logger.warning(f"Quantity: {len(false_keys)}")

                true_keys = [key for key in checker_dict.keys() if key not in false_keys]
                logger.debug(f"Open a true key for reference: {true_keys[0]}")

                # test_name = self.TestOptions.get()
                # batch_num = self.BatchOptions.get() # Batch 1
                # batch_num = int(batch_num.split(" ")[1]) # 1
                # treatment_dir = Path(THE_HISTORY.get_treatment_dir(self.CURRENT_PROJECT, 
                #                                                     test_name, 
                #                                                     batch_num, 
                #                                                     treatment_char))
                project_dir = Path(THE_HISTORY.get_project_dir(self.CURRENT_PROJECT))
                
                if "Shoaling" in true_keys[0]:
                    ref_path = project_dir / true_keys[0] / "trajectories_nogaps.txt"
                else:
                    ref_path = project_dir / true_keys[0] / "trajectories.txt"
                ref_df, ref_tanks = load_raw_df(ref_path)

                skip_list = []
                for key in false_keys:
                    _message = f"Trajectory {key} is invalid."
                    _message += f"\n Press YES, in the case of FREEZING fish, you are required to input a valid coordinate for it"
                    _message += f"\n Press NO, in the case of DEAD/NO fish, which means its end-points would be left blank in the final result."
                    _message += f"\n Press CANCEL, if there is supposed to be a valid trajectory for this fish/tank."
                    choice = tkinter.messagebox.askyesnocancel("Warning", _message)
                    if choice == True:
                        logger.debug(f"User chose to fill coordinates for {key}")
                        if "Shoaling" in key:
                            filler_path = project_dir / key / "trajectories_nogaps.txt"
                        else:
                            filler_path = project_dir / key / "trajectories.txt"

                        inputted = self.trajectories_filler(filler_path, ref_df, ref_tanks)

                        if inputted == False:
                            skip_num = Path(key).stem
                            skip_list.append(skip_num)
                            continue
                    elif choice == False:
                        logger.debug(f"User chose to skip {key}")
                        skip_num = Path(key).stem
                        skip_list.append(skip_num)
                        continue
                    else:
                        logger.debug(f"User chose to cancel")
                        return False
                
                skip_dict[treatment_char] = skip_list

            return skip_dict

        return True

    def analyze_project_THREADED(self):
        logger.debug("Open a new thread to analyze project")

        status = self.pre_analyze_check()
        if status == False:
            logger.info("Pre-analyze check failed, aborting analyze project")
            return
        elif status == True:
            self.SKIP_DICT = {}
        else:
            self.SKIP_DICT = status

        try:
            # Close all self.checker_windows
            for window in self.checker_windows:
                window.destroy()
        except Exception as e:
            logger.warning(e)

        analyze_thread = threading.Thread(target=self.analyze_project)
        analyze_thread.start()


    ### IMPORT BUTTON ###

    def import_trajectories(self):
        logger.debug("Import trajectories button pressed")

        if self.CURRENT_PROJECT == "":
            tkinter.messagebox.showerror("Error", "Please select a project")
            return
        
        project_dir = Path(THE_HISTORY.get_project_dir(self.CURRENT_PROJECT))

        # Get the project directory from the user using a file dialog
        ori_dir = tkinter.filedialog.askdirectory()
        print("Selected directory:", ori_dir)

        # Convert the project directory to a Path object
        if ori_dir:
            ori_dir = Path(ori_dir)
        else:
            return

        # Find all .txt files within the project directory
        txt_files = ori_dir.glob("**/*.txt")

        # change all paths in txt_files to fullpath
        txt_paths = [txt_file.resolve() for txt_file in txt_files]

        char_list = [chr(i) for i in range(65, 91)]

        def check_grandparent_format(txt_path, batch_num):
            # check if the grandparent folder of txt_path is in the format of "A - Control (1st Batch)"
            # if yes, return True, else return False

            grandparent = txt_path.parent.parent.name
            sign = grandparent.split("-")[0].strip()
            treatment = grandparent.split("-")[1].strip()
            batch_ord = treatment.split("(")[1].split(" ")[0].strip()
            # change ordinal to number, 1st -> 1
            if batch_num != int(batch_ord[:-2]):
                return None
            if sign not in char_list:
                # change it from number to letter, 1 -> A
                sign = chr(int(sign) + 64)
            
            return f"{sign} - {treatment}"

        def check_test(txt_path):
            # check if any word in name_dict.keys() is in str(txt_path).lower()
            # if yes then construct a name based on the found name
            name_dict = {
                "novel tank" : "01 - Novel Tank Test",
                "shoaling" : "02 - Shoaling Test",
                "mirror" : "03 - Mirror Biting Test",
                "social" : "04 - Social Interaction Test",
                "predator" : "05 - Predator Test"
            }

            for key in name_dict.keys():
                if key in str(txt_path).lower():
                    return name_dict[key]

        def get_project_path(txt_path, project_dir, batch_num):
            gparents = check_grandparent_format(txt_path, batch_num)
            if gparents == None:
                return None
            # if the txt_path contain "novel tank"
            ancestors = Path(check_test(txt_path)) / Path(gparents)
            # get the parent and name of txt_path
            parent = txt_path.parent.name
            file = txt_path.name
            return project_dir / ancestors / parent / file

        BATCH_NUMBER = int(self.BatchOptions.get().split()[1])

        # Loop through each .txt file and copy it to the new location
        for txt_path in txt_paths:
            # Construct the new path by replacing the project directory with the new base directory
            new_path = get_project_path(txt_path, project_dir, BATCH_NUMBER)

            if new_path == None:
                continue
            
            # Create the new directory structure if it doesn't already exist
            new_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Copy the file to the new location
            shutil.copy(txt_path, new_path)
            
            print("Copied {} to {}".format(txt_path, new_path))

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)
        logger.debug(f"New UI appearance mode: {new_appearance_mode}")

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        logger.debug(f"New UI scaling: {new_scaling_float}")
        customtkinter.set_widget_scaling(new_scaling_float)


if __name__ == "__main__":
    app = App()
    app.mainloop()